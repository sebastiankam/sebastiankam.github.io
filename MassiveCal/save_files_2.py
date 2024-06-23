# -*- coding: utf-8 -*-
"""Save Files

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1ac6wUIsPlgKZtVqYGaiybY_84pfSqIjx
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Source workbook name
source_workbook_name = "Temp.xlsx"

# Get the current date and time
current_datetime = datetime.now()
datetime_string = current_datetime.strftime("%Y%m%d%H%M")

# Destination workbook name
destination_workbook_name = f"Backup Log - {datetime_string}.xlsx"
save_directory = "Backup Log"

# Create the save directory if it doesn't exist
os.makedirs(save_directory, exist_ok=True)

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_workbook_name, data_only=True)

# Create a new workbook for backup
backup_workbook = openpyxl.Workbook()
default_sheet = backup_workbook.active
backup_workbook.remove(default_sheet)

# Copy values from source workbook to backup workbook
for sheet_name in ["Test1", "Test2", "Test3", "Test4", "Test5"]:
    source_sheet = source_workbook[sheet_name]
    backup_sheet = backup_workbook.create_sheet(title=sheet_name)

    for row in source_sheet.iter_rows(values_only=True):
        backup_sheet.append(row)

# Save the backup workbook
backup_workbook.save(os.path.join(save_directory, destination_workbook_name))

import openpyxl

# Existing workbook name
existing_workbook_name = "Temp.xlsx"

# Load the existing workbook
existing_workbook = openpyxl.load_workbook(existing_workbook_name)

# Remove existing sheets named "Test1" to "Test5"
for sheet_name in ["Test1", "Test2", "Test3", "Test4", "Test5"]:
    if sheet_name in existing_workbook.sheetnames:
        existing_workbook.remove(existing_workbook[sheet_name])

# Create new sheets named "Test1" to "Test5"
for sheet_name in ["Test1", "Test2", "Test3", "Test4", "Test5"]:
    existing_workbook.create_sheet(title=sheet_name)

# Save the modified workbook
existing_workbook.save(existing_workbook_name)

"""# 新增區段"""

Option Explicit

Sub CopySheetsAndPasteValues()
    Dim sourceWorkbook As Workbook
    Dim newWorkbook As Workbook
    Dim sourceSheet As Worksheet
    Dim newSheet As Worksheet
    Dim dateTimeString As String
    Dim saveDirectory As String

    ' Set the source workbook
    Set sourceWorkbook = ThisWorkbook

    ' Get the current date and time
    dateTimeString = Format(Now(), "YYYYMMDDHHMM")

    ' Get the saving directory of the current workbook
    saveDirectory = sourceWorkbook.Path

    ' Create a new workbook
    Set newWorkbook = Workbooks.Add

    ' Copy and paste values for sheets "Test1", "Test2", "Test3"
    For Each sourceSheet In sourceWorkbook.Sheets
        If sourceSheet.Name = "Test1" Or sourceSheet.Name = "Test2" Or sourceSheet.Name = "Test3" Then
            Set newSheet = newWorkbook.Sheets.Add(After:=newWorkbook.Sheets(newWorkbook.Sheets.Count))
            sourceSheet.UsedRange.Copy
            newSheet.Range("A1").PasteSpecial (xlPasteFormulasAndNumberFormats)
            Application.CutCopyMode = False
            'newSheet.Range("A1").PasteSpecial (xlPasteValuesAndNumberFormats)

            ' Modify formulas to remove external references
            Dim cell As Range
            For Each cell In newSheet.UsedRange.SpecialCells(xlCellTypeFormulas)
                cell.Formula = Replace(cell.Formula, sourceWorkbook.Name, newWorkbook.Name)
            Next cell

            newSheet.Name = sourceSheet.Name
        End If
    Next sourceSheet

    ' Save the new workbook in the same directory with the specified name
    newWorkbook.SaveAs saveDirectory & "\Data - " & dateTimeString & ".xlsx"

    ' Clean up
    newWorkbook.Close
    Set newWorkbook = Nothing
    Set sourceWorkbook = Nothing
End Sub

import pandas as pd

# Create a DataFrame
df = pd.DataFrame(columns=['Data'])

# Create a list to be saved in the DataFrame cell
my_list = [1, 2, 3, 4, 5]

# Save the list in a DataFrame cell
df.loc[0, 'Data'] = my_list

# Access the list from the DataFrame cell
retrieved_list = df.loc[0, 'Data']

print(df)

Option Explicit

Sub CopySheetsAndPasteValues()
    Dim sourceWorkbook As Workbook
    Dim newWorkbook As Workbook
    Dim sourceSheet As Worksheet
    Dim newSheet As Worksheet
    Dim dateTimeString As String
    Dim saveDirectory As String

    ' Set the source workbook
    Set sourceWorkbook = ThisWorkbook

    ' Get the current date and time
    dateTimeString = Format(Now(), "YYYYMMDDHHMM")

    ' Get the saving directory of the current workbook
    saveDirectory = sourceWorkbook.Path

    ' Create a new workbook
    Set newWorkbook = Workbooks.Add

    ' Copy and paste values for sheets "Test1", "Test2", "Test3"
    For Each sourceSheet In sourceWorkbook.Sheets
        If sourceSheet.Name = "Test1" Or sourceSheet.Name = "Test2" Or sourceSheet.Name = "Test3" Then
            Set newSheet = newWorkbook.Sheets.Add(After:=newWorkbook.Sheets(newWorkbook.Sheets.Count))
            sourceSheet.UsedRange.Copy
            newSheet.Range("A1").PasteSpecial (xlPasteFormulasAndNumberFormats)
            Application.CutCopyMode = False
            'newSheet.Range("A1").PasteSpecial (xlPasteValuesAndNumberFormats)

            ' Modify formulas to remove external references
            Dim cell As Range
            For Each cell In newSheet.UsedRange.SpecialCells(xlCellTypeFormulas)
                cell.Formula = Replace(cell.Formula, sourceWorkbook.Name, newWorkbook.Name)
            Next cell

            newSheet.Name = sourceSheet.Name
        End If
    Next sourceSheet

    ' Save the new workbook in the same directory with the specified name
    newWorkbook.SaveAs saveDirectory & "\Data - " & dateTimeString & ".xlsx"

    ' Clean up
    newWorkbook.Close
    Set newWorkbook = Nothing
    Set sourceWorkbook = Nothing
End Sub

Sub SelectAndPasteFilePath()
    Dim filePath As String
    Dim targetCell As Range

    ' Open file dialog window to select Excel file
    With Application.FileDialog(msoFileDialogFilePicker)
        .Title = "Select Excel File"
        .Filters.Clear
        .Filters.Add "Excel Files", "*.xls; *.xlsx; *.xlsm"
        .AllowMultiSelect = False
        If .Show = -1 Then
            ' Get the selected file path
            filePath = .SelectedItems(1)

            ' Set the target cell where the file path will be pasted
            Set targetCell = ActiveSheet.Range("Z10") ' Modify as per your requirement

            ' Paste the file path in the target cell
            targetCell.Value = filePath
        End If
    End With
End Sub

Sub ReadDataWorkbook()
    Dim dataFilePath As String
    Dim dataWorkbook As Workbook
    Dim dataSheet As Worksheet
    Dim itemName As String
    Dim regionName As String
    Dim sheetName As String
    Dim headerRow As Range
    Dim dataRange As Range
    Dim targetSheet As Worksheet
    Dim targetCell As Range

    ' Get the data file path from Cell Z10
    dataFilePath = Range("Z10").Value

    ' Check if the data file path is provided
    If dataFilePath = "" Then
        MsgBox "Data file path is not specified in Cell Z10."
        Exit Sub
    End If

    ' Get the item name, region name, and sheet name from respective cells
    itemName = Range("B1").Value
    regionName = Range("B2").Value
    sheetName = Range("B3").Value

    ' Check if either item name or region name is provided
    If itemName = "" And regionName = "" Then
        MsgBox "Either item name (Cell B1) or region name (Cell B2) must be specified."
        Exit Sub
    End If

    ' Open the data workbook
    On Error Resume Next
    Set dataWorkbook = Workbooks.Open(dataFilePath)
    On Error GoTo 0

    ' Check if the data workbook was successfully opened
    If dataWorkbook Is Nothing Then
        MsgBox "Unable to open the data workbook. Please verify the file path in Cell Z10."
        Exit Sub
    End If

    ' Set the data sheet based on the specified sheet name
    On Error Resume Next
    Set dataSheet = dataWorkbook.Sheets(sheetName)
    On Error GoTo 0

    ' Check if the data sheet exists
    If dataSheet Is Nothing Then
        MsgBox "The specified sheet name does not exist in the data workbook."
        dataWorkbook.Close False
        Exit Sub
    End If

    ' Set the header row and data range
    Set headerRow = dataSheet.Range("A1").CurrentRegion.Rows(1)
    Set dataRange = dataSheet.Range("A2").CurrentRegion

    ' Set the target worksheet and cell for pasting the extracted data
    Set targetSheet = ThisWorkbook.ActiveSheet ' Modify if necessary
    Set targetCell = targetSheet.Range("B10") ' Modify if necessary

    ' Perform the operations based on the provided cell values
    If itemName <> "" Then
        ' Find the row with the specified item name
        Dim itemRow As Range
        Set itemRow = dataRange.Columns(1).Find(itemName, LookIn:=xlValues, LookAt:=xlWhole)

        ' Check if the item name is found
        If itemRow Is Nothing Then
            MsgBox "Item name '" & itemName & "' not found in the data workbook."
        Else
            ' Copy the header row and the row with the item name to the target worksheet
            headerRow.Copy Destination:=targetCell
            dataRange.Rows(itemRow.Row - dataRange.Row + 1).Copy Destination:=targetCell.Offset(1)
        End If
    ElseIf regionName <> "" Then
        ' Filter the data range based on the region name
        dataRange.AutoFilter Field:=2, Criteria1:=regionName

        ' Copy the filtered data (including header row) to the target worksheet
        dataRange.SpecialCells(xlCellTypeVisible).Copy Destination:=targetCell

        ' Turn off the filter in the data range
        dataSheet.AutoFilterMode = False
    End If

    ' Close the data workbook without saving changes
    'dataWorkbook.Close False

    ' Clean up
    Set dataSheet = Nothing
    Set dataWorkbook = Nothing
    Set targetCell = Nothing
    Set targetSheet = Nothing

    MsgBox "Data extraction completed successfully."
End Sub