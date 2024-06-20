# -*- coding: utf-8 -*-
"""MassiveCalculation

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1275QTwx63nNkCvhywRLSk5cJFEdXMB8E
"""

import openpyxl
import pandas as pd
import numpy as np

# Load the Excel file
workbook = openpyxl.load_workbook("Temp - VBA.xlsm")

# Read the "Hi" worksheet
hi_sheet = workbook["Hi"]
company_col = pd.DataFrame([cell.value for cell in hi_sheet["A"]], columns=["Company"]).dropna()
time_col = pd.DataFrame([cell.value for cell in hi_sheet["N"]], columns=["TimeSet"]).dropna()

print(company_col)
print(time_col)

# Read the "Data" worksheet
data_sheet = workbook["Data"]
data2_sheet = workbook["Data2"]

# Initialize the benchmark dictionary
benchmark = {}

# Iterate over the rows in the benchmark data range
for row in data2_sheet.iter_rows(values_only=True):
    item_name = row[0]
    test_b_value = row[1]
    test_c_value = row[2]
    benchmark[item_name] = (test_b_value, test_c_value)

if "TestA" in benchmark:
    del benchmark["TestA"]

benchmark

# Create initial empty dataframes

def clear_charts():
    global more_item, less_item, more_item_value, less_item_value
    column_names = time_col["TimeSet"].tolist()
    more_item = pd.DataFrame(columns=["ItemName"] + column_names)
    less_item = pd.DataFrame(columns=["ItemName"] + column_names)
    more_item_value = pd.DataFrame(columns=["ItemName"] + column_names)
    less_item_value = pd.DataFrame(columns=["ItemName"] + column_names)

clear_charts()

# Create a list to store the data
data = []

# Iterate through the rows in the "Data" worksheet
for row in data_sheet.iter_rows(values_only=True):
    data.append(row)

# Create the DataFrame
data_df = pd.DataFrame(data)

# Set the column names (if available)
if data_sheet.max_row > 0:
    data_df.columns = data_sheet[1]

# Display the DataFrame
data_df.columns = range(data_df.shape[1])
data_df

# Get processing company and time, delete extra columns, rename
def get_block(data_df, row_extract_start, row_extract_end, col_extract_start, col_extract_end):
    block_df = data_df.iloc[row_extract_start:row_extract_end, col_extract_start:col_extract_end]
    processing_company = block_df.iloc[0, 3]
    processing_time = block_df.iloc[1, 3]
    block_df = block_df.drop(block_df.columns[3], axis=1)
    block_df = block_df.dropna(how='all', axis=0)
    block_df.columns = block_df.iloc[0]
    block_df = block_df.iloc[1:].reset_index(drop=True)
    return block_df, processing_company, processing_time

# Check if the item is in the more_item list already. Add if not.
def check_or_add_item(item_name):
    global more_item, less_item, more_item_value, less_item_value
    if not more_item.empty:
        if item_name in more_item["ItemName"].values:
            return
    new_row = pd.DataFrame({"ItemName": [item_name]})
    more_item = pd.concat([more_item, new_row]).fillna(0)
    less_item = pd.concat([less_item, new_row]).fillna(0)
    more_item_value = pd.concat([more_item_value, new_row]).fillna(0)
    less_item_value = pd.concat([less_item_value, new_row]).fillna(0)

# Clear previous records
clear_charts()

# Slice the data_df slowly
row_extract_start = 0
row_extract_end = 10
col_extract_start = 0
col_extract_end = 4

while row_extract_end <= len(data_df)+5:

    while col_extract_end <= len(data_df.columns):
        block_df, processing_company, processing_time = get_block(data_df, row_extract_start, row_extract_end, col_extract_start, col_extract_end)
        diff = 0

        for _, row in block_df.iterrows():

            # Check item added in four charts or not; also for dictionary
            check_or_add_item(row["TestA"])
            if row["TestA"] not in benchmark:
                benchmark.setdefault(row["TestA"], (0, 0))

           # Compare with benchmark
            if row["TestA"] not in benchmark:
                diff = row["TestB"]
            else:
                diff = row["TestB"] - benchmark[row["TestA"]][0]

            # Update the dataframes
            if diff > 0:
                more_item.loc[more_item["ItemName"] == row["TestA"], processing_time] += 1
                more_item_value.loc[more_item_value["ItemName"] == row["TestA"], processing_time] += row["TestC"] - benchmark[row["TestA"]][1]
            elif diff < 0:
                less_item.loc[less_item["ItemName"] == row["TestA"], processing_time] += 1
                less_item_value.loc[less_item_value["ItemName"] == row["TestA"], processing_time] += benchmark[row["TestA"]][1] - row["TestC"]
            else:
                pass

        col_extract_start += 5
        col_extract_end += 5

        print(processing_company, processing_time)

    row_extract_start += 11
    row_extract_end += 11
    col_extract_start = 0
    col_extract_end = 4

more_item

less_item

more_item_value

less_item_value

from openpyxl.utils.dataframe import dataframe_to_rows

# Create worksheets and set their names
worksheet_names = ["WS1", "WS2", "WS3", "WS4"]
worksheets = [workbook.create_sheet(title=name) for name in worksheet_names]

# Get references to the existing worksheets
#worksheet_names = ["WS1", "WS2", "WS3", "WS4"]
#worksheets = [workbook[name] for name in worksheet_names]

# Define the data to be written to worksheets
data = [
    more_item,
    less_item,
    more_item_value,
    less_item_value
]

# Write data to worksheets
for sheet, value in zip(worksheets, data):
    column_names = value.columns.tolist()
    sheet.append(column_names)  # Write column names in the first row
    for row in dataframe_to_rows(value, index=False, header=False):
        sheet.append(row)

# Save the workbook
workbook.save("Temp.xlsx")

